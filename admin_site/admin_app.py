from __future__ import annotations

import csv
import datetime as dt
import io
import json
import logging
import os
import sys
# Fix shared import path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from functools import wraps

from flask import Flask, flash, jsonify, redirect, render_template, request, Response, send_file, send_from_directory, session, url_for
from flask_mail import Mail, Message
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from PIL import Image


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from shared import *

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder="templates", static_folder="static", static_url_path="/admin/static")
app.config.from_object('shared.config.Config')
mail = Mail(app)

# init_db() moved to __main__ only


# ── Security Headers ──
@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if os.environ.get("FLASK_ENV", "development").lower() == "production":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


# ── Error Handlers ──
@app.errorhandler(404)
def not_found(error):
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(error):
    logger.error("Server error: %s", error, exc_info=True)
    return render_template("500.html"), 500


# ── Health Check ──
@app.route("/health")
def health_check():
    return jsonify({"status": "ok", "service": "busresort-admin"}), 200



def _get_room_pricing_map(conn) -> dict[str, float]:
    rows = conn.execute("SELECT room_type, price_per_night FROM RoomPricing").fetchall()
    return {r["room_type"]: float(r["price_per_night"]) for r in rows}


def _get_room_pricing_rows(conn):
    rows = conn.execute("SELECT id, room_type, price_per_night FROM RoomPricing ORDER BY id").fetchall()
    return [dict(row) for row in rows]


def _get_resort_rooms(conn):
    return conn.execute(
        "SELECT id, name, room_type, capacity, COALESCE(status, 'Available') AS status, "
        "COALESCE(image_path, '') AS image_path "
        "FROM ResortRooms ORDER BY room_type, name"
    ).fetchall()


def _get_resort_room_photos(conn, room_id: int):
    return conn.execute(
        """
        SELECT photo_order, image_path
        FROM ResortRoomPhotos
        WHERE room_id = ?
        ORDER BY photo_order ASC
        """,
        (room_id,),
    ).fetchall()






def _get_bus_pricing(conn):
    return conn.execute("SELECT destination, price FROM BusPricing ORDER BY destination").fetchall()


def _get_exclusive_price(conn) -> float:
    row = conn.execute("SELECT exclusive_price FROM ResortOptions WHERE id = 1").fetchone()
    return float(row["exclusive_price"]) if row else 0.0


def _get_rentable_appliances(conn):
    return conn.execute("SELECT id, name, price FROM RentableAppliances ORDER BY name").fetchall()


def _parse_date(value: str) -> dt.date | None:
    try:
        return dt.date.fromisoformat(value)
    except Exception:
        return None


def _normalize_time(value: str, default: str) -> str:
    candidate = (value or "").strip()
    try:
        return dt.time.fromisoformat(candidate).strftime("%H:%M")
    except Exception:
        return default


def _parse_datetime(value: str) -> dt.datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return dt.datetime.fromisoformat(raw)
    except Exception:
        return None


def _row_value(row, key: str, default=None):
    """Safely access sqlite3.Row values with a fallback default."""
    if row is None:
        return default
    try:
        value = row[key]
    except Exception:
        return default
    return default if value is None else value


def _get_date_range(filter_type: str, start_date_str: str = "", end_date_str: str = ""):
    """Return (start_date, end_date, label) for a given filter type.
    Dates are ISO strings (YYYY-MM-DD) or None for all-time.
    """
    today = dt.date.today()
    ft = (filter_type or "").strip().lower()

    if ft == "daily":
        d = today.strftime("%Y-%m-%d")
        return d, d, f"Daily — {today.strftime('%B %d, %Y')}"

    if ft == "weekly":
        # Monday-based week
        monday = today - dt.timedelta(days=today.weekday())
        sunday = monday + dt.timedelta(days=6)
        return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d"), f"Weekly — {monday.strftime('%b %d')} to {sunday.strftime('%b %d, %Y')}"

    if ft == "monthly":
        start = today.replace(day=1)
        # Last day of month
        if today.month == 12:
            end = today.replace(day=31)
        else:
            end = (today.replace(month=today.month + 1, day=1)) - dt.timedelta(days=1)
        return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"), f"Monthly — {today.strftime('%B %Y')}"

    if ft == "yearly":
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31)
        return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"), f"Yearly — {today.strftime('%Y')}"

    if ft == "custom":
        s = _parse_date(start_date_str)
        e = _parse_date(end_date_str)
        if s and e:
            if s > e:
                s, e = e, s
            return s.strftime("%Y-%m-%d"), e.strftime("%Y-%m-%d"), f"Custom — {s.strftime('%b %d, %Y')} to {e.strftime('%b %d, %Y')}"
        if s:
            return s.strftime("%Y-%m-%d"), None, f"Custom — From {s.strftime('%b %d, %Y')}"
        if e:
            return None, e.strftime("%Y-%m-%d"), f"Custom — Until {e.strftime('%b %d, %Y')}"

    return None, None, "All Time"


def _ensure_upload_folder() -> str:
    """Get the unified uploads folder path from config."""
    folder = app.config.get("UPLOAD_FOLDER")
    if not folder:
        # Fallback to project root static/uploads
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        folder = os.path.join(project_root, "static", "uploads")
    os.makedirs(folder, exist_ok=True)
    return folder


ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5MB limit

# Room gallery constraints
ROOM_PHOTO_MAX = 3
ROOM_PHOTO_ASPECT_W = 4
ROOM_PHOTO_ASPECT_H = 5


def _save_processed_room_photo(file_storage, slot_index: int) -> str | None:
    """Save a processed 4:5 portrait room photo and return the URL path.

    Processing behavior:
      - Validates extension already handled upstream.
      - Converts to RGB (for webp/png/jpg consistently).
      - Crops to 4:5 centered (main subject centered as requested).
      - Maintains high quality.

    Returns:
      - '/uploads/<filename>' relative URL path.
    """
    if not file_storage or not file_storage.filename:
        return None

    upload_dir = _ensure_upload_folder()
    raw_name = secure_filename(file_storage.filename)
    if not raw_name:
        return None

    stem, ext = os.path.splitext(raw_name)
    ext_lower = ext.lower()
    if ext_lower not in ALLOWED_IMAGE_EXTENSIONS:
        return None

    # Use a stable naming scheme to simplify debugging; order is captured in filename
    timestamp = dt.datetime.now().strftime('%Y%m%d%H%M%S%f')
    out_name = f"room_{slot_index}_{timestamp}_{stem}{ext_lower}"
    out_full_path = os.path.join(upload_dir, out_name)

    try:
        # Pillow needs the stream from the beginning
        file_storage.seek(0)
        img = Image.open(file_storage)
        img = img.convert('RGB')

        w, h = img.size
        target_ratio = ROOM_PHOTO_ASPECT_W / ROOM_PHOTO_ASPECT_H  # 0.8
        current_ratio = w / h if h else target_ratio

        # Determine crop rectangle centered
        if current_ratio > target_ratio:
            # Image is too wide -> crop width
            new_w = int(h * target_ratio)
            left = (w - new_w) // 2
            box = (left, 0, left + new_w, h)
        else:
            # Image is too tall/narrow -> crop height
            new_h = int(w / target_ratio)
            top = (h - new_h) // 2
            box = (0, top, w, top + new_h)

        img_cropped = img.crop(box)

        # Export as 4:5 with same cropped content; keep resolution for quality
        img_cropped.save(out_full_path, quality=95, optimize=True)

        return f"/uploads/{out_name}"
    except Exception as e:
        logger.error(f"Failed to process room photo: {e}")
        return None



def _save_website_image(file_storage, image_type: str = "image") -> str | None:
    """Save website content image and return relative path.
    
    Args:
        file_storage: Werkzeug FileStorage object
        image_type: Type of image (homepage, resort, bus) for naming
    
    Returns:
        Relative path like 'uploads/filename.jpg' or None on failure
    """
    if not file_storage or not file_storage.filename:
        return None
    
    # Use unified upload folder from config
    upload_dir = _ensure_upload_folder()
    
    # Use secure_filename for the original filename
    raw_name = secure_filename(file_storage.filename)
    if not raw_name:
        return None
    
    # Validate file extension (jpg, jpeg, png, webp only)
    stem, ext = os.path.splitext(raw_name)
    ext_lower = ext.lower()
    if ext_lower not in ALLOWED_IMAGE_EXTENSIONS:
        logger.warning(f"Invalid image extension: {ext}")
        return None
    
    # Generate unique filename with timestamp and image type prefix
    timestamp = dt.datetime.now().strftime('%Y%m%d%H%M%S%f')
    unique_name = f"{image_type}_{timestamp}_{stem}{ext_lower}"
    save_path = os.path.join(upload_dir, unique_name)
    
    try:
        file_storage.save(save_path)
        logger.info(f"Website image saved: {save_path}")
        # Return relative path for the URL (uploads/filename.jpg)
        return f"uploads/{unique_name}"
    except Exception as e:
        logger.error(f"Failed to save website image: {e}")
        return None


def _delete_image_file(image_path: str) -> bool:
    """Delete an image file from the uploads folder."""
    if not image_path:
        return False
    try:
        # Extract filename from path
        filename = os.path.basename(image_path)
        if not filename:
            return False
        
        # Build full path
        upload_dir = _ensure_upload_folder()
        full_path = os.path.join(upload_dir, filename)
        
        # Check if file exists and delete it
        if os.path.exists(full_path):
            os.remove(full_path)
            logger.info(f"Deleted image file: {full_path}")
            return True
        return False
    except Exception as e:
        logger.error(f"Failed to delete image file {image_path}: {e}")
        return False


def _save_uploaded_room_image(file_storage) -> str | None:
    """Save uploaded room image and return the URL path."""
    if not file_storage or not file_storage.filename:
        return None
    
    upload_dir = _ensure_upload_folder()
    raw_name = secure_filename(file_storage.filename)
    if not raw_name:
        return None
    
    # Validate file extension
    stem, ext = os.path.splitext(raw_name)
    ext_lower = ext.lower()
    if ext_lower not in ALLOWED_IMAGE_EXTENSIONS:
        logger.warning(f"Invalid image extension: {ext}")
        return None
    
    # Generate unique filename
    unique_name = f"room_{dt.datetime.now().strftime('%Y%m%d%H%M%S%f')}_{stem}{ext_lower}"
    save_path = os.path.join(upload_dir, unique_name)
    
    try:
        file_storage.save(save_path)
        logger.info(f"Room image saved: {save_path}")
        # Return relative URL path for /uploads/ route
        return f"/uploads/{unique_name}"
    except Exception as e:
        logger.error(f"Failed to save room image: {e}")
        return None


def _send_email(to_email: str, subject: str, body: str) -> tuple[bool, str]:
    if not (app.config.get("MAIL_SERVER") and app.config.get("MAIL_USERNAME") and app.config.get("MAIL_PASSWORD")):
        return False, "Email is not configured (set MAIL_SERVER/MAIL_USERNAME/MAIL_PASSWORD)."
    try:
        msg = Message(subject=subject, recipients=[to_email], body=body)
        mail.send(msg)
        return True, "Email sent."
    except Exception as e:
        return False, str(e)


@app.context_processor
def inject_site_settings():
    """Inject site settings into all templates - safe with default values if table is missing or DB error."""
    conn = None
    try:
        conn = get_db_connection(timeout=5.0)
        settings = conn.execute("SELECT * FROM WebsiteSettings WHERE id = 1").fetchone()
        conn.close()
        return {"site_settings": settings}
    except Exception as e:
        logger.warning(f"Could not load site settings (non-fatal): {e}")
        if conn:
            try:
                conn.close()
            except:
                pass
        # Return None rather than crashing - templates handle missing settings gracefully
        return {"site_settings": None}


@app.context_processor
def inject_admin_notifications():
    if session.get('admin_id'):
        conn = None
        try:
            conn = get_db_connection(timeout=5.0)  # Short timeout for context processor
            unread_count, notifications = get_notifications(conn)
            return {"unread_count": unread_count, "notifications": notifications}
        except Exception as e:
            logger.error(f"Failed to load notifications: {e}")
            return {"unread_count": 0, "notifications": []}
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
    return {"unread_count": 0, "notifications": []}


def admin_login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("admin_id"):
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)
    return wrapped


@app.route("/")
def admin_root():
    import logging
    from shared import get_db_connection
    conn = get_db_connection()
    logging.getLogger(__name__).info("[ROUTE ENTRY] /admin entry conn_type=%s get_db_connection_from=%s", type(conn).__name__, get_db_connection.__module__)
    conn.close()
    return redirect(url_for("admin_login"))


@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    logger.info(f"[LOGIN] Request method: {request.method}")
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        logger.info(f"[LOGIN] Attempt for user: {username}")

        conn = get_db_connection()
        admin = conn.execute(
            "SELECT id, username, password FROM Admin WHERE username = ?",
            (username,),
        ).fetchone()
        conn.close()

        if admin and check_password_hash(admin["password"], password):
            logger.info(f"[LOGIN] Password match successful for {username}")
            session.clear()
            session["admin_id"] = admin["id"]
            session["admin_username"] = admin["username"]
            session.modified = True
            logger.info(f"[LOGIN] Session set - admin_id: {session.get('admin_id')}")
            logger.info(f"[LOGIN] Redirecting to dashboard...")
            resp = redirect(url_for("dashboard"), code=302)
            logger.info(f"[LOGIN] Response created - Location: {resp.headers.get('Location')}")
            return resp
        else:
            logger.warning(f"[LOGIN] Failed for user: {username} - Admin found: {admin is not None}")

        flash("Invalid admin credentials.", "error")

    return render_template("admin_login.html")


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/api/admin/stats")
@admin_login_required
def api_admin_stats():
    conn = None
    try:
        conn = get_db_connection(timeout=10.0)
        active_bus = conn.execute("SELECT COUNT(*) as cnt FROM BusBookings WHERE status IN ('Pending','Confirmed')").fetchone()['cnt']
        active_resort = conn.execute("SELECT COUNT(*) as cnt FROM ResortBookings WHERE status IN ('Pending','Confirmed')").fetchone()['cnt']
        total_bookings = int(active_bus) + int(active_resort)
        active_reservations = total_bookings

        total_customers = conn.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM (
                SELECT email FROM BusBookings
                WHERE status IN ('Pending','Confirmed')
                  AND email IS NOT NULL AND email != ''
                UNION
                SELECT email FROM ResortBookings
                WHERE status IN ('Pending','Confirmed')
                  AND email IS NOT NULL AND email != ''
            ) t
            """
        ).fetchone()['cnt']

        bus_revenue = conn.execute("SELECT COALESCE(SUM(price),0) as total FROM BusBookings WHERE status='Confirmed'").fetchone()['total']
        resort_revenue = conn.execute("SELECT COALESCE(SUM(price),0) as total FROM ResortBookings WHERE status='Confirmed'").fetchone()['total']
        total_revenue = float(bus_revenue) + float(resort_revenue)

        return jsonify({
            'total_bookings': total_bookings,
            'active_reservations': active_reservations,
            'total_customers': total_customers,
            'total_revenue': total_revenue,
        }), 200
    except Exception as e:
        logger.error(f"API stats error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to load stats'}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route("/dashboard")
@admin_login_required
def dashboard():
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        appliances = _get_rentable_appliances(conn)
        room_pricing_rows = _get_room_pricing_rows(conn)
        resort_rooms = _get_resort_rooms(conn)
        # Attach ordered photo list for rendering previews
        resort_rooms_list = []
        for room in resort_rooms:
            room_dict = dict(room)
            photos = _get_resort_room_photos(conn, room_dict["id"])
            room_dict["photos"] = [p["image_path"] for p in photos if p["image_path"]]
            resort_rooms_list.append(room_dict)

        exclusive_price = _get_exclusive_price(conn)

        # Notification retrieval can fail if DB is partially initialized on Vercel.
        # Root-cause: real exception should be fixed upstream, but we must not
        # crash the entire dashboard when Notification table/columns are missing.
        # If Notification is absent, we fall back to safe defaults.
        try:
            unread_count, notifications = get_notifications(conn)
        except Exception as notif_exc:
            logger.error("Dashboard notification load failed (non-fatal): %s", notif_exc, exc_info=True)
            unread_count, notifications = 0, []

        # Summary stats (ACTIVE/VALID only; never include Cancelled)

        # Active statuses: Pending + Confirmed
        active_bus = conn.execute(
            "SELECT COUNT(*) as cnt FROM BusBookings WHERE status IN ('Pending','Confirmed')"
        ).fetchone()['cnt']
        active_resort = conn.execute(
            "SELECT COUNT(*) as cnt FROM ResortBookings WHERE status IN ('Pending','Confirmed')"
        ).fetchone()['cnt']

        total_bus = active_bus
        total_resort = active_resort
        total_bookings = total_bus + total_resort
        active_reservations = total_bookings

        # Revenue/cards should exclude cancelled entirely. Confirmed only.
        bus_revenue = conn.execute(
            "SELECT COALESCE(SUM(price),0) as total FROM BusBookings WHERE status='Confirmed'"
        ).fetchone()['total']
        resort_revenue = conn.execute(
            "SELECT COALESCE(SUM(price),0) as total FROM ResortBookings WHERE status='Confirmed'"
        ).fetchone()['total']
        total_revenue = float(bus_revenue) + float(resort_revenue)

        # Total customers: only customers with at least one ACTIVE or COMPLETED booking.
        # Current schema uses 'Confirmed' as completed; 'Pending' as active.
        total_customers = conn.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM (
                SELECT email FROM BusBookings
                WHERE status IN ('Pending','Confirmed')
                  AND email IS NOT NULL AND email != ''
                UNION
                SELECT email FROM ResortBookings
                WHERE status IN ('Pending','Confirmed')
                  AND email IS NOT NULL AND email != ''
            ) t
            """
        ).fetchone()['cnt']

        return render_template(

            "dashboard.html",
            appliances=appliances,
            room_pricing_rows=room_pricing_rows,
            resort_rooms=resort_rooms_list,

            exclusive_price=exclusive_price,
            unread_count=unread_count,
            notifications=notifications,
            total_bookings=total_bookings,
            active_reservations=active_reservations,
            total_revenue=total_revenue,
            total_customers=total_customers,
        )
    except Exception as e:
        logger.error(f"Dashboard error: {e}", exc_info=True)
        flash("An error occurred while loading the dashboard. Please try again.", "error")
        return render_template("dashboard.html", 
            appliances=[],
            room_pricing_rows=[],
            resort_rooms=[],
            exclusive_price=0.0,
            unread_count=0, notifications=[],
            total_bookings=0, active_reservations=0, total_revenue=0, total_customers=0
        ), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/admin/requests/bus")
@admin_login_required
def admin_bus_requests():
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        bus_bookings = conn.execute("SELECT * FROM BusBookings ORDER BY id DESC").fetchall()
        return render_template("admin_bus_requests.html", bus_bookings=bus_bookings)
    except Exception as e:
        logger.error(f"Bus requests error: {e}", exc_info=True)
        flash("Error loading bus bookings. Please refresh the page.", "error")
        return render_template("admin_bus_requests.html", bus_bookings=[]), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/admin/requests/resort")
@admin_login_required
def admin_resort_requests():
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        resort_bookings = conn.execute("SELECT * FROM ResortBookings ORDER BY id DESC").fetchall()
        return render_template("admin_resort_requests.html", resort_bookings=resort_bookings)
    except Exception as e:
        logger.error(f"Resort requests error: {e}", exc_info=True)
        flash("Error loading resort bookings. Please refresh the page.", "error")
        return render_template("admin_resort_requests.html", resort_bookings=[]), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/confirm/<booking_type>/<int:booking_id>", methods=["POST"])
@admin_login_required
def confirm_booking(booking_type: str, booking_id: int):
    table = "BusBookings" if booking_type == "bus" else "ResortBookings" if booking_type == "resort" else None
    if not table:
        flash("Invalid booking type.", "error")
        return redirect(url_for("dashboard"))

    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        
        # Check if booking exists first
        booking = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (booking_id,)).fetchone()
        if not booking:
            flash("Booking not found.", "error")
            return redirect(request.referrer or url_for("dashboard"))
        
        # Check if already confirmed
        if booking["status"] == "Confirmed":
            flash("This booking is already confirmed.", "warning")
            return redirect(request.referrer or url_for("dashboard"))
        
        # Update status
        conn.execute(f"UPDATE {table} SET status = 'Confirmed' WHERE id = ?", (booking_id,))
        conn.commit()
        
        # Send email notification (don't crash if email fails)
        booking_email = _row_value(booking, "email", "")
        if booking_email:
            try:
                if booking_type == "bus":
                    details = (
                        f"Pickup: {_row_value(booking, 'pickup', 'N/A')}\n"
                        f"Destination: {_row_value(booking, 'destination', 'N/A')}\n"
                        f"Date/Time: {_row_value(booking, 'datetime', 'N/A')}\n"
                    )
                else:
                    details = (
                        f"Check-in: {_row_value(booking, 'checkin', 'N/A')}\n"
                        f"Check-out: {_row_value(booking, 'checkout', 'N/A')}\n"
                        f"Check-in time: {_row_value(booking, 'checkin_time', '14:00')}\n"
                        f"Check-out time: {_row_value(booking, 'checkout_time', '12:00')}\n"
                        f"Guests: {_row_value(booking, 'guests', 'N/A')}\n"
                        f"Room Type: {_row_value(booking, 'room_type', 'N/A')}\n"
                    )
                body = (
                    f"Hello {_row_value(booking, 'name', 'Customer')},\n\n"
                    f"Your {booking_type} booking has been CONFIRMED.\n\n"
                    f"Booking details:\n{details}\n"
                    f"Thank you."
                )
                _send_email(booking_email, "Booking Confirmation", body)
            except Exception as email_err:
                # Log email error but don't crash the booking confirmation
                logger.warning(f"Failed to send confirmation email for booking {booking_id}: {email_err}")
        
        # Create notification
        try:
            notif_type = f'booking_{booking_type}'
            notif_msg = f"New {booking_type} booking confirmed for {_row_value(booking, 'name', 'Unknown')}"
            safe_create_notification(conn, notif_msg, notif_type)
        except Exception as notif_err:
            logger.warning(f"Failed to create notification: {notif_err}")
        
        flash("Booking confirmed successfully.", "success")
        
    except Exception as exc:
        # Postgres (psycopg2) and SQLite (dev fallback) may raise different exception types.
        # Keep a single handler to avoid relying on sqlite3-specific exception classes.
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error(
            f"Database error confirming {booking_type} booking {booking_id}: {exc}",
            exc_info=True,
        )
        flash("A database error occurred. Please try again.", "error")

    except Exception as exc:
        if conn:
            conn.rollback()
        logger.error(f"Error confirming {booking_type} booking {booking_id}: {exc}", exc_info=True)
        flash(f"An error occurred while confirming the booking: {str(exc)}", "error")
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass
    
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/cancel/<booking_type>/<int:booking_id>", methods=["POST"])
@admin_login_required
def cancel_booking(booking_type: str, booking_id: int):
    """Cancel a booking (bus or resort).

    Fixes root-cause server errors by avoiding name collision with the shared
    cancellation helper, and by using a single explicit transaction.

    If called via AJAX (fetch/XHR), returns JSON.
    """
    import traceback as _traceback

    table = "BusBookings" if booking_type == "bus" else "ResortBookings" if booking_type == "resort" else None
    date_column = "datetime" if booking_type == "bus" else "checkin"

    def _is_ajax() -> bool:
        return request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json or request.accept_mimetypes.best == "application/json"

    if not table:
        msg = "Invalid booking type."
        if _is_ajax():
            return jsonify({"success": False, "error": msg}), 400
        flash(msg, "error")
        return redirect(url_for("dashboard"))

    conn = None
    try:
        conn = get_db_connection(timeout=30.0)

        booking = conn.execute(
            f"SELECT id, name, status, {date_column} FROM {table} WHERE id = ?",
            (booking_id,),
        ).fetchone()

        if not booking:
            msg = "Booking not found."
            if _is_ajax():
                return jsonify({"success": False, "error": msg}), 404
            flash(msg, "error")
            return redirect(request.referrer or url_for("dashboard"))

        if booking["status"] == "Cancelled":
            msg = "This booking has already been cancelled."
            if _is_ajax():
                return jsonify({"success": False, "error": msg, "already_cancelled": True}), 409
            flash(msg, "warning")
            return redirect(request.referrer or url_for("dashboard"))

        # Use shared helper without name collision.
        from shared import db as shared_db
        cancel_fn = getattr(shared_db, "cancel_booking")

        # Transaction handling
        conn.execute("BEGIN")
        res = cancel_fn(conn, booking_type, booking_id, "admin")
        if not res.get("found"):
            raise RuntimeError("Cancellation helper could not find the booking during update.")
        if res.get("already_cancelled"):
            # Treat as idempotent success-ish
            pass

        conn.commit()

        # Success response
        success_msg = "Booking cancelled successfully."
        if _is_ajax():
            return jsonify({
                "success": True,
                "message": success_msg,
                "booking_type": booking_type,
                "booking_id": booking_id,
                "status": "Cancelled",
            }), 200

        flash(f"{booking_type.title()} booking cancelled successfully.", "success")
        return redirect(request.referrer or url_for("dashboard"))

    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        # Log complete traceback to terminal
        logger.error("Error cancelling %s booking %s: %s", booking_type, booking_id, exc, exc_info=True)
        _traceback.print_exc()

        err_msg = "An unexpected error occurred while cancelling the booking."
        if _is_ajax():
            return jsonify({"success": False, "error": err_msg, "details": str(exc)}), 500
        flash(err_msg, "error")
        return redirect(request.referrer or url_for("dashboard"))

    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass



@app.route("/delete/<booking_type>/<int:booking_id>", methods=["POST"])
@admin_login_required
def delete_booking(booking_type: str, booking_id: int):
    table = "BusBookings" if booking_type == "bus" else "ResortBookings" if booking_type == "resort" else None
    if not table:
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    conn.execute(f"DELETE FROM {table} WHERE id = ?", (booking_id,))
    conn.commit()
    conn.close()
    flash("Booking deleted.", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/edit/<booking_type>/<int:booking_id>", methods=["GET", "POST"])
@admin_login_required
def edit_booking(booking_type: str, booking_id: int):
    table = "BusBookings" if booking_type == "bus" else "ResortBookings" if booking_type == "resort" else None
    if not table:
        flash("Invalid booking type.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    room_type_options = [r["room_type"] for r in _get_room_pricing_rows(conn)]
    booking = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (booking_id,)).fetchone()
    if not booking:
        conn.close()
        flash("Booking not found.", "error")
        return redirect(url_for("dashboard"))
    booking_data = dict(booking)

    if request.method == "POST":
        form = request.form
        if booking_type == "bus":
            price = float(form.get("price", booking_data["price"] or 0) or 0)
            conn.execute(
                """
                UPDATE BusBookings
                SET name = ?, contact = ?, email = ?, pickup = ?, destination = ?, datetime = ?, checkin = ?, checkout = ?, status = ?, price = ?
                WHERE id = ?
                """,
                (
                    form.get("name", "").strip(),
                    form.get("contact", "").strip(),
                    form.get("email", "").strip(),
                    form.get("pickup", "").strip(),
                    form.get("destination", "").strip(),
                    form.get("datetime", "").strip(),
                    form.get("datetime", "").strip(),
                    form.get("checkout", "").strip(),
                    form.get("status", "Pending").strip() or "Pending",
                    price,
                    booking_id,
                ),
            )
        else:
            payment_method = (form.get("payment_method", booking_data.get("payment_method", "Cash") or "Cash") or "Cash").strip()
            if payment_method not in ("Cash", "GCash", "Bank Transfer"):
                payment_method = "Cash"
            price_per_night = float(form.get("price_per_night", booking_data.get("price_per_night", 0) or 0) or 0)
            total_cost = float(form.get("total_cost", booking_data.get("total_cost", booking_data["price"] or 0) or 0) or 0)
            conn.execute(
                """
                UPDATE ResortBookings
                SET name = ?, contact = ?, email = ?, checkin = ?, checkout = ?, checkin_time = ?, checkout_time = ?, guests = ?, room_type = ?, payment_method = ?, status = ?, price_per_night = ?, total_cost = ?, price = ?
                WHERE id = ?
                """,
                (
                    form.get("name", "").strip(),
                    form.get("contact", "").strip(),
                    form.get("email", "").strip(),
                    form.get("checkin", "").strip(),
                    form.get("checkout", "").strip(),
                    _normalize_time(form.get("checkin_time", booking_data.get("checkin_time", "14:00")), "14:00"),
                    _normalize_time(form.get("checkout_time", booking_data.get("checkout_time", "12:00")), "12:00"),
                    int(form.get("guests", "0") or 0),
                    form.get("room_type", "").strip(),
                    payment_method,
                    form.get("status", "Pending").strip() or "Pending",
                    price_per_night,
                    total_cost,
                    total_cost,
                    booking_id,
                ),
            )
        conn.commit()
        conn.close()
        flash("Booking updated.", "success")
        return redirect(url_for("dashboard"))

    return render_template(
        "edit_booking.html",
        booking_type=booking_type,
        booking=booking_data,
        room_type_options=room_type_options,
    )


@app.route("/admin/bus-pricing", methods=["POST"])
@admin_login_required
def admin_update_bus_pricing():
    destination = (request.form.get("destination", "") or "").strip()
    raw_price = request.form.get("price", "0")
    try:
        price = float(raw_price or 0)
    except ValueError:
        price = 0.0

    if not destination:
        flash("Destination is required.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    conn.execute(
        "INSERT INTO BusPricing (destination, price) VALUES (?, ?) "
        "ON CONFLICT(destination) DO UPDATE SET price = excluded.price",
        (destination, price),
    )
    conn.commit()
    conn.close()
    flash("Bus destination pricing saved.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing", methods=["POST"])
@admin_login_required
def admin_update_resort_pricing():
    conn = get_db_connection()
    for room_type in ("Standard", "Deluxe", "Family"):
        raw = request.form.get(room_type, "0")
        try:
            price = float(raw or 0)
        except ValueError:
            price = 0.0
        conn.execute(
            "INSERT INTO RoomPricing (room_type, price_per_night) VALUES (?, ?) "
            "ON CONFLICT(room_type) DO UPDATE SET price_per_night = excluded.price_per_night",
            (room_type, price),
        )
    try:
        exclusive_price = float(request.form.get("exclusive_price", "0") or 0)
    except ValueError:
        exclusive_price = 0.0
    conn.execute(
        "INSERT INTO ResortOptions (id, exclusive_price) VALUES (1, ?) "
        "ON CONFLICT(id) DO UPDATE SET exclusive_price = excluded.exclusive_price",
        (exclusive_price,),
    )
    conn.commit()
    conn.close()
    flash("Resort pricing updated.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/add", methods=["POST"])
@admin_login_required
def admin_add_resort_price_option():
    room_type = (request.form.get("room_type", "") or "").strip()
    try:
        price = float(request.form.get("price_per_night", "0") or 0)
    except ValueError:
        price = 0.0
    if not room_type:
        flash("Room type is required.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO RoomPricing (room_type, price_per_night) VALUES (?, ?) "
        "ON CONFLICT(room_type) DO UPDATE SET price_per_night = excluded.price_per_night",
        (room_type, price),
    )
    conn.commit()
    conn.close()
    flash("Resort pricing option added.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/edit-by-id/<int:pricing_id>", methods=["POST"])
@admin_login_required
def admin_edit_resort_price_option_by_id(pricing_id: int):
    room_type = (request.form.get("room_type", "") or "").strip()
    try:
        price = float(request.form.get("price_per_night", "0") or 0)
    except ValueError:
        price = 0.0
    if not room_type:
        flash("Room type is required.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db_connection()
    current = conn.execute("SELECT room_type FROM RoomPricing WHERE id = ?", (pricing_id,)).fetchone()
    if not current:
        conn.close()
        flash("Pricing option not found.", "error")
        return redirect(url_for("dashboard"))
    duplicate = conn.execute(
        "SELECT id FROM RoomPricing WHERE room_type = ? AND id <> ?",
        (room_type, pricing_id),
    ).fetchone()
    if duplicate:
        conn.close()
        flash("That room type already exists.", "error")
        return redirect(url_for("dashboard"))
    conn.execute(
        "UPDATE RoomPricing SET room_type = ?, price_per_night = ? WHERE id = ?",
        (room_type, price, pricing_id),
    )
    if current["room_type"] != room_type:
        conn.execute(
            "UPDATE ResortRooms SET room_type = ? WHERE room_type = ?",
            (room_type, current["room_type"]),
        )
    conn.commit()
    conn.close()
    flash("Resort pricing option updated.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/delete-by-id/<int:pricing_id>", methods=["POST"])
@admin_login_required
def admin_delete_resort_price_option_by_id(pricing_id: int):
    conn = get_db_connection()
    option = conn.execute("SELECT room_type FROM RoomPricing WHERE id = ?", (pricing_id,)).fetchone()
    if not option:
        conn.close()
        flash("Pricing option not found.", "error")
        return redirect(url_for("dashboard"))
    rooms_using = conn.execute(
        "SELECT COUNT(1) AS cnt FROM ResortRooms WHERE room_type = ?",
        (option["room_type"],),
    ).fetchone()
    if rooms_using and int(rooms_using["cnt"] or 0) > 0:
        conn.close()
        flash("Cannot delete pricing option while rooms still use this type.", "error")
        return redirect(url_for("dashboard"))
    conn.execute("DELETE FROM RoomPricing WHERE id = ?", (pricing_id,))
    conn.commit()
    conn.close()
    flash("Resort pricing option deleted.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/delete/<room_type>", methods=["POST"])
@admin_login_required
def admin_delete_resort_room_price(room_type: str):
    if room_type not in ("Standard", "Deluxe", "Family"):
        flash("Invalid room type.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db_connection()
    conn.execute("DELETE FROM RoomPricing WHERE room_type = ?", (room_type,))
    conn.commit()
    conn.close()
    flash(f"{room_type} pricing option deleted.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/edit/<room_type>", methods=["POST"])
@admin_login_required
def admin_edit_resort_room_price(room_type: str):
    if room_type not in ("Standard", "Deluxe", "Family"):
        flash("Invalid room type.", "error")
        return redirect(url_for("dashboard"))
    try:
        new_price = float(request.form.get("price", "0") or 0)
    except ValueError:
        new_price = 0.0
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO RoomPricing (room_type, price_per_night) VALUES (?, ?) "
        "ON CONFLICT(room_type) DO UPDATE SET price_per_night = excluded.price_per_night",
        (room_type, new_price),
    )
    conn.commit()
    conn.close()
    flash(f"{room_type} pricing updated.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/delete-exclusive", methods=["POST"])
@admin_login_required
def admin_delete_exclusive_resort_price():
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO ResortOptions (id, exclusive_price) VALUES (1, 0) "
        "ON CONFLICT(id) DO UPDATE SET exclusive_price = 0"
    )
    conn.commit()
    conn.close()
    flash("Exclusive resort pricing option deleted.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/resort-pricing/edit-exclusive", methods=["POST"])
@admin_login_required
def admin_edit_exclusive_resort_price():
    try:
        exclusive_price = float(request.form.get("exclusive_price", "0") or 0)
    except ValueError:
        exclusive_price = 0.0
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO ResortOptions (id, exclusive_price) VALUES (1, ?) "
        "ON CONFLICT(id) DO UPDATE SET exclusive_price = excluded.exclusive_price",
        (exclusive_price,),
    )
    conn.commit()
    conn.close()
    flash("Exclusive resort pricing updated.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/rooms/add", methods=["POST"])
@admin_login_required
def admin_add_room():
    name = request.form.get("name", "").strip()
    room_type = request.form.get("room_type", "").strip()
    try:
        capacity = int(request.form.get("capacity", 2) or 2)
    except ValueError:
        capacity = 2
    if not name or not room_type:
        flash("Name and room type required.", "error")
        return redirect(url_for("dashboard"))

    # Handle up to 3 room photos (new gallery)
    photo_paths: list[str] = ["", "", ""]
    uploaded_any = False

    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT OR IGNORE INTO RoomPricing (room_type, price_per_night) VALUES (?, 0)",
            (room_type,),
        )

        # Determine if room exists by name
        existing = conn.execute(
            "SELECT id, COALESCE(image_path, '') AS image_path FROM ResortRooms WHERE name = ?",
            (name,),
        ).fetchone()

        room_id = int(existing["id"]) if existing else None

        # If room exists, we keep current ResortRooms.image_path for backwards compat, but photos go to ResortRoomPhotos.
        # Process uploads first (fail fast on invalid size/format)
        for slot_index in range(3):
            file_field = f"photo_{slot_index}"
            file_storage = request.files.get(file_field)
            if not file_storage or not file_storage.filename:
                continue

            uploaded_any = True

            # Check file size (5MB limit)
            file_storage.seek(0, os.SEEK_END)
            file_size = file_storage.tell()
            file_storage.seek(0)
            if file_size > MAX_IMAGE_SIZE:
                flash("Each photo must be <= 5MB.", "error")
                conn.close()
                return redirect(url_for("dashboard"))

            # Process to 4:5 portrait
            processed_path = _save_processed_room_photo(file_storage, slot_index)
            if not processed_path:
                flash("Failed to process photo. Use JPG, PNG, WEBP.", "error")
                conn.close()
                return redirect(url_for("dashboard"))

            photo_paths[slot_index] = processed_path

        if existing:
            conn.execute(
                "UPDATE ResortRooms SET room_type = ?, capacity = ?, status = COALESCE(status, 'Available') WHERE id = ?",
                (room_type, capacity, existing["id"]),
            )
            room_id = int(existing["id"])
            msg = f"Room '{name}' updated."
        else:
            # Keep legacy single image_path for now: slot 0 if provided.
            legacy_image = photo_paths[0] if photo_paths[0] else ""
            cursor = conn.execute(
                """
                INSERT INTO ResortRooms (name, room_type, capacity, image_path)
                VALUES (?, ?, ?, ?)
                RETURNING id
                """,
                (name, room_type, capacity, legacy_image),
            )
            room_id = cursor.fetchone()["id"]
            msg = f"Room '{name}' added."


        # Persist gallery photos to ResortRoomPhotos (ordered 0..2)
        for slot_index, path in enumerate(photo_paths):
            if not path:
                continue
            conn.execute(
                """
                INSERT INTO ResortRoomPhotos (room_id, photo_order, image_path)
                VALUES (?, ?, ?)
                ON CONFLICT(room_id, photo_order) DO UPDATE SET image_path = excluded.image_path
                """,
                (room_id, slot_index, path),
            )

        conn.commit()
        flash(msg + (" Photo(s) uploaded." if uploaded_any else ""), "success")
        return redirect(url_for("dashboard"))
    finally:
        try:
            conn.close()
        except:
            pass



@app.route("/admin/rooms/edit/<int:room_id>", methods=["POST"])
@admin_login_required
def admin_edit_room(room_id: int):
    name = (request.form.get("name", "") or "").strip()
    room_type = (request.form.get("room_type", "") or "").strip()
    raw_status = (request.form.get("status", "Available") or "Available").strip()
    status = raw_status if raw_status in ("Available", "Unavailable") else "Available"
    try:
        capacity = int(request.form.get("capacity", "2") or 2)
    except ValueError:
        capacity = 2

    if not name or not room_type:
        flash("Room name and room type are required.", "error")
        return redirect(url_for("dashboard"))
    if capacity < 1:
        flash("Capacity must be at least 1.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    conn.execute(
        "INSERT OR IGNORE INTO RoomPricing (room_type, price_per_night) VALUES (?, 0)",
        (room_type,),
    )
    existing = conn.execute(
        "SELECT id, COALESCE(image_path, '') AS image_path FROM ResortRooms WHERE id = ?",
        (room_id,),
    ).fetchone()
    if not existing:
        conn.close()
        flash("Room not found.", "error")
        return redirect(url_for("dashboard"))

    duplicate = conn.execute(
        "SELECT id FROM ResortRooms WHERE name = ? AND id <> ?",
        (name, room_id),
    ).fetchone()
    if duplicate:
        conn.close()
        flash("A different room already uses that name.", "error")
        return redirect(url_for("dashboard"))

    # Handle up to 3 room photos (new gallery). Existing single image_path is kept for backward compat.

    photo_paths: list[str] = ["", "", ""]
    uploaded_any = False

    # process optional uploads
    for slot_index in range(3):
        file_field = f"photo_{slot_index}"
        image_file = request.files.get(file_field)
        if not image_file or not image_file.filename:
            continue

        uploaded_any = True

        # Check file size (5MB limit)
        image_file.seek(0, os.SEEK_END)
        file_size = image_file.tell()
        image_file.seek(0)

        if file_size > MAX_IMAGE_SIZE:
            flash("Each photo must be <= 5MB.", "error")
            conn.close()
            return redirect(url_for("dashboard"))

        processed_path = _save_processed_room_photo(image_file, slot_index)
        if not processed_path:
            flash("Failed to process photo. Use JPG, PNG, WEBP.", "error")
            conn.close()
            return redirect(url_for("dashboard"))

        photo_paths[slot_index] = processed_path

    # Update base room fields (do NOT break legacy image_path behavior)
    conn.execute(
        "UPDATE ResortRooms SET name = ?, room_type = ?, capacity = ?, status = ? WHERE id = ?",
        (name, room_type, capacity, status, room_id),
    )

    # Persist/replace gallery photos in ResortRoomPhotos
    if uploaded_any:
        for slot_index, path in enumerate(photo_paths):
            if not path:
                continue
            conn.execute(
                """
                INSERT INTO ResortRoomPhotos (room_id, photo_order, image_path)
                VALUES (?, ?, ?)
                ON CONFLICT(room_id, photo_order) DO UPDATE SET image_path = excluded.image_path
                """,
                (room_id, slot_index, path),
            )

    conn.commit()
    conn.close()

    flash("Room updated." + (" Photo(s) updated." if uploaded_any else ""), "success")
    return redirect(url_for("dashboard"))



@app.route("/admin/rooms/toggle-status/<int:room_id>", methods=["POST"])
@admin_login_required
def admin_toggle_room_status(room_id):
    conn = get_db_connection()
    room = conn.execute("SELECT status FROM ResortRooms WHERE id = ?", (room_id,)).fetchone()
    if room:
        new_status = 'Unavailable' if room['status'] == 'Available' else 'Available'
        conn.execute("UPDATE ResortRooms SET status = ? WHERE id = ?", (new_status, room_id))
        conn.commit()
        flash(f"Room status changed to {new_status}.", "success")
    conn.close()
    return redirect(url_for("dashboard"))


@app.route("/admin/rooms/delete/<int:room_id>", methods=["POST"])
@admin_login_required
def admin_delete_room(room_id):
    conn = get_db_connection()
    room = conn.execute("SELECT name, image_path FROM ResortRooms WHERE id = ?", (room_id,)).fetchone()
    if room:
        # Delete associated image file if exists
        if room['image_path']:
            _delete_image_file(room['image_path'])
        
        conn.execute("DELETE FROM ResortRooms WHERE id = ?", (room_id,))
        conn.commit()
        flash(f"Room '{room['name']}' deleted.", "success")
    conn.close()
    return redirect(url_for("dashboard"))


@app.route("/admin/rooms/remove-photo/<int:room_id>/<int:slot_index>", methods=["POST"])
@admin_login_required
def admin_remove_room_photo(room_id: int, slot_index: int):
    """Remove a single gallery photo slot (0..2) from a room."""
    if slot_index not in (0, 1, 2):
        flash("Invalid photo slot.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT image_path FROM ResortRoomPhotos WHERE room_id = ? AND photo_order = ?",
            (room_id, slot_index),
        ).fetchone()

        if row and row["image_path"]:
            _delete_image_file(row["image_path"])

        conn.execute(
            "DELETE FROM ResortRoomPhotos WHERE room_id = ? AND photo_order = ?",
            (room_id, slot_index),
        )

        # Maintain legacy image_path for slot 0
        if slot_index == 0:
            new_path = conn.execute(
                "SELECT image_path FROM ResortRoomPhotos WHERE room_id = ? ORDER BY photo_order ASC LIMIT 1",
                (room_id,),
            ).fetchone()
            legacy = new_path["image_path"] if new_path else ""
            conn.execute("UPDATE ResortRooms SET image_path = ? WHERE id = ?", (legacy, room_id))

        conn.commit()
        flash("Photo removed.", "success")
    finally:
        conn.close()

    return redirect(url_for("dashboard"))



@app.route("/admin/appliances/add", methods=["POST"])
@admin_login_required
def admin_add_appliance():
    name = (request.form.get("name", "") or "").strip()
    try:
        price = float(request.form.get("price", "0") or 0)
    except ValueError:
        price = 0.0
    if not name:
        flash("Appliance name is required.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO RentableAppliances (name, price) VALUES (?, ?) "
        "ON CONFLICT(name) DO UPDATE SET price = excluded.price",
        (name, price),
    )
    conn.commit()
    conn.close()
    flash("Appliance saved.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/appliances/delete/<int:appliance_id>", methods=["POST"])
@admin_login_required
def admin_delete_appliance(appliance_id: int):
    conn = get_db_connection()
    conn.execute("DELETE FROM RentableAppliances WHERE id = ?", (appliance_id,))
    conn.commit()
    conn.close()
    flash("Appliance removed.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/appliances/edit/<int:appliance_id>", methods=["POST"])
@admin_login_required
def admin_edit_appliance(appliance_id: int):
    name = (request.form.get("name", "") or "").strip()
    try:
        price = float(request.form.get("price", "0") or 0)
    except ValueError:
        price = 0.0
    if not name:
        flash("Appliance name is required.", "error")
        return redirect(url_for("dashboard"))
    conn = get_db_connection()
    conn.execute(
        "UPDATE RentableAppliances SET name = ?, price = ? WHERE id = ?",
        (name, price, appliance_id),
    )
    conn.commit()
    conn.close()
    flash("Appliance updated.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/content", methods=["GET", "POST"])
@admin_login_required
def admin_content():
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        settings, contact, footer = _load_site_content_records(conn)

        if request.method == "POST":
            errors, upload_errors = _save_site_content(conn, settings, request.form, request.files)
            if errors:
                for e in errors:
                    flash(e, "error")
                conn.rollback()
                return render_template("admin_content.html", settings=settings, contact=contact, footer=footer), 400

            if upload_errors:
                flash("Some images failed to upload. Please try again.", "warning")
            else:
                flash("Site content updated successfully.", "success")

            settings, contact, footer = _load_site_content_records(conn)

        return render_template("admin_content.html", settings=settings, contact=contact, footer=footer)

    except Exception as e:
        import traceback
        logger.error(f"Admin content error: {e}", exc_info=True)
        traceback.print_exc()
        flash(f"Error loading/saving site content: {str(e)}", "error")
        return render_template(
            "admin_content.html",
            settings=settings if 'settings' in locals() else None,
            contact=contact if 'contact' in locals() else None,
            footer=footer if 'footer' in locals() else None,
        ), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass




def _sqlite_row_to_dict(row):
    return dict(row) if row is not None else None


def _load_site_content_records(conn):
    """Load WebsiteSettings + CMS rows, ensuring schema and default rows exist."""
    migrate_cms_schema(conn)
    settings = conn.execute("SELECT * FROM WebsiteSettings WHERE id = 1").fetchone()
    contact = conn.execute("SELECT * FROM CMS_ContactInfo WHERE id = 1").fetchone()
    footer = conn.execute("SELECT * FROM CMS_FooterContent WHERE id = 1").fetchone()
    if contact is None:
        conn.execute(
            """
            INSERT INTO CMS_ContactInfo (id)
            VALUES (1)
            ON CONFLICT (id) DO NOTHING
            """
        )
        contact = conn.execute("SELECT * FROM CMS_ContactInfo WHERE id = 1").fetchone()
    if footer is None:
        conn.execute(
            """
            INSERT INTO CMS_FooterContent (id)
            VALUES (1)
            ON CONFLICT (id) DO NOTHING
            """
        )
        footer = conn.execute("SELECT * FROM CMS_FooterContent WHERE id = 1").fetchone()

    return settings, contact, footer


def _site_content_field(data, key, default=""):
    if hasattr(data, "get"):
        val = data.get(key, default)
    else:
        val = default
    return (val or default).strip() if isinstance(val, str) else (val or default)


def _save_site_content(conn, settings, data, files=None):
    """Persist Site Content from a form-like mapping. Returns (errors, upload_errors)."""
    import re

    def _validate_email(email: str) -> bool:
        email = (email or "").strip()
        return bool(email and re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email))

    def _validate_phone(phone: str) -> bool:
        phone = (phone or "").strip()
        if not phone:
            return False
        for ch in phone:
            if ch.isdigit() or ch in "+() -.":
                continue
            return False
        return len([c for c in phone if c.isdigit()]) >= 7

    def _validate_url(url: str) -> bool:
        url = (url or "").strip()
        return not url or url.startswith("http://") or url.startswith("https://")

    files = files or {}
    site_name = _site_content_field(data, "site_name", "BusResort") or "BusResort"
    homepage_welcome = _site_content_field(data, "homepage_welcome")
    homepage_description = _site_content_field(data, "homepage_description")
    contact_email = _site_content_field(data, "contact_email")

    # Booking success popup fields (stored in CMS_Homepage)
    popup_title = _site_content_field(data, "booking_success_popup_title", "Booking Submitted Successfully!")
    popup_message = _site_content_field(data, "booking_success_popup_message", "Thank you for choosing BusResort!\nYour booking has been successfully submitted and is now awaiting confirmation from our administrator. We appreciate your trust in our service and look forward to serving you. Please keep your booking reference number for future inquiries.")

    popup_show_icon = 1 if (data.get("booking_success_popup_show_icon") in ["on", "1", 1, True, "true"]) else 0


    homepage_image = settings["homepage_image"] if settings else ""
    resort_image = settings["resort_image"] if settings else ""
    bus_image = settings["bus_image"] if settings else ""
    upload_errors = []

    for field, kind in [("homepage_image", "homepage"), ("resort_image", "resort"), ("bus_image", "bus")]:
        file = files.get(field) if hasattr(files, "get") else None
        if file and getattr(file, "filename", None):
            result = _save_website_image(file, kind)
            if result:
                if field == "homepage_image":
                    homepage_image = result
                elif field == "resort_image":
                    resort_image = result
                else:
                    bus_image = result
            else:
                upload_errors.append(kind)

    errors = []
    if not site_name:
        errors.append("Site name is required.")
    if not contact_email or not _validate_email(contact_email):
        errors.append("A valid Contact Email is required.")

    business_name = _site_content_field(data, "business_name")
    business_tagline = _site_content_field(data, "business_tagline")
    email = _site_content_field(data, "email")
    mobile = _site_content_field(data, "mobile")
    google_maps_link = _site_content_field(data, "google_maps_link")
    facebook_url = _site_content_field(data, "facebook_url")
    instagram_url = _site_content_field(data, "instagram_url")
    x_url = _site_content_field(data, "x_url")
    tiktok_url = _site_content_field(data, "tiktok_url")
    footer_description = _site_content_field(data, "footer_description")
    contact_section_title = _site_content_field(data, "contact_section_title")
    copyright_text = _site_content_field(data, "copyright_text")
    terms_href = _site_content_field(data, "terms_href")

    if not business_name:
        errors.append("Business Name is required.")
    if not business_tagline:
        errors.append("Business Description is required.")
    if not email or not _validate_email(email):
        errors.append("Valid Contact Email is required.")
    if not mobile or not _validate_phone(mobile):
        errors.append("Valid Mobile Number is required.")
    if not google_maps_link or not _validate_url(google_maps_link):
        errors.append("Valid Google Maps URL is required.")

    for label, u in [
        ("Facebook", facebook_url),
        ("Instagram", instagram_url),
        ("X", x_url),
        ("TikTok", tiktok_url),
        ("Terms & Conditions", terms_href),
    ]:
        if u and not _validate_url(u):
            errors.append(f"{label} URL must start with http:// or https://")

    if errors:
        return errors, upload_errors

    conn.execute(
        """UPDATE WebsiteSettings SET
            site_name = ?, homepage_welcome = ?, homepage_description = ?, contact_email = ?,
            homepage_image = ?, resort_image = ?, bus_image = ?
            WHERE id = 1""",
        (site_name, homepage_welcome, homepage_description, contact_email,
         homepage_image, resort_image, bus_image),
    )
    conn.execute(
        """UPDATE CMS_ContactInfo SET
            business_name = ?, business_tagline = ?, mobile = ?, email = ?,
            google_maps_link = ?, facebook_url = ?, instagram_url = ?, x_url = ?, tiktok_url = ?
            WHERE id = 1""",
        (business_name, business_tagline, mobile, email, google_maps_link,
         facebook_url, instagram_url, x_url, tiktok_url),
    )
    conn.execute(
        """UPDATE CMS_FooterContent SET
            footer_description = ?, contact_section_title = ?, copyright_text = ?, terms_href = ?
            WHERE id = 1""",
        (footer_description, contact_section_title, copyright_text, terms_href),
    )

    # Persist booking success popup content (buttons are hardcoded on the frontend)
    conn.execute(
        """UPDATE CMS_Homepage SET
            booking_success_popup_title = ?,
            booking_success_popup_message = ?,
            booking_success_popup_show_icon = ?
            WHERE id = 1""",
        (popup_title, popup_message, popup_show_icon),
    )

    conn.execute("DELETE FROM CMS_SocialLinks WHERE platform IN ('facebook','instagram','x','tiktok')")
    for i, (platform, href) in enumerate([
        ("facebook", facebook_url),
        ("instagram", instagram_url),
        ("x", x_url),
        ("tiktok", tiktok_url),
    ]):
        href = (href or "").strip()
        if href:
            conn.execute(
                "INSERT INTO CMS_SocialLinks (platform, href, is_visible, link_order) VALUES (?, ?, 1, ?)",
                (platform, href, i),
            )
    conn.commit()
    return [], upload_errors


@app.route("/api/admin/site-content", methods=["GET", "POST"])
@admin_login_required
def api_site_content():
    """JSON API for loading and saving Site Content."""
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)
        settings, contact, footer = _load_site_content_records(conn)

        if request.method == "GET":
            return jsonify({
                "success": True,
                "settings": _sqlite_row_to_dict(settings),
                "contact": _sqlite_row_to_dict(contact),
                "footer": _sqlite_row_to_dict(footer),
            }), 200

        payload = request.get_json(silent=True) if request.is_json else request.form
        errors, upload_errors = _save_site_content(conn, settings, payload, request.files)
        if errors:
            return jsonify({"success": False, "error": errors[0], "errors": errors}), 400

        settings, contact, footer = _load_site_content_records(conn)
        message = "Site content updated successfully."
        if upload_errors:
            message = "Site content saved, but some images failed to upload."
        return jsonify({
            "success": True,
            "message": message,
            "settings": _sqlite_row_to_dict(settings),
            "contact": _sqlite_row_to_dict(contact),
            "footer": _sqlite_row_to_dict(footer),
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"API site content error: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route("/admin/content/remove-image/<image_type>", methods=["POST"])
@admin_login_required
def admin_remove_website_image(image_type):
    """Remove a website content image (homepage, resort, or bus)."""
    valid_types = {'homepage_image', 'resort_image', 'bus_image'}
    if image_type not in valid_types:
        flash("Invalid image type.", "error")
        return redirect(url_for("admin_content"))
    
    conn = get_db_connection()
    settings = conn.execute("SELECT * FROM WebsiteSettings WHERE id = 1").fetchone()
    
    if settings and settings[image_type]:
        # Delete the image file
        _delete_image_file(settings[image_type])
        
        # Update database to remove image path
        conn.execute(f"UPDATE WebsiteSettings SET {image_type} = '' WHERE id = 1")
        conn.commit()
        flash(f"{image_type.replace('_', ' ').title()} removed.", "success")
    else:
        flash("No image to remove.", "warning")
    
    conn.close()
    return redirect(url_for("admin_content"))


def _admin_email_changed(conn, admin_id: int, new_email: str) -> bool:
    row = conn.execute("SELECT email FROM Admin WHERE id = ?", (admin_id,)).fetchone()
    old_email = (row["email"] or "").strip() if row else ""
    return old_email != (new_email or "").strip()


def send_admin_notification(subject: str, message: str) -> tuple[bool, str]:
    conn = get_db_connection()
    admin = conn.execute("SELECT email, email_enabled FROM Admin LIMIT 1").fetchone()
    conn.close()

    if not admin or not admin["email"] or admin["email_enabled"] == 0:
        return False, "Email notifications are not enabled or admin email is not set."

    if not (app.config.get("MAIL_SERVER") and app.config.get("MAIL_USERNAME") and app.config.get("MAIL_PASSWORD")):
        return False, "Email is not configured (set MAIL_SERVER/MAIL_USERNAME/MAIL_PASSWORD)."

    try:
        msg = Message(subject=subject, recipients=[admin["email"]], body=message)
        mail.send(msg)
        return True, "Email sent."
    except Exception as e:
        return False, str(e)


@app.route("/admin/contact-info", methods=["GET", "POST"])
@admin_login_required
def admin_contact_info():
    conn = None
    try:
        conn = get_db_connection(timeout=30.0)

        contact = conn.execute("SELECT * FROM CMS_ContactInfo WHERE id = 1").fetchone()
        footer = conn.execute("SELECT * FROM CMS_FooterContent WHERE id = 1").fetchone()

        if contact is None:
            conn.execute(
                """
                INSERT INTO CMS_ContactInfo (id)
                VALUES (1)
                ON CONFLICT (id) DO NOTHING
                """
            )
            contact = conn.execute("SELECT * FROM CMS_ContactInfo WHERE id = 1").fetchone()

        if footer is None:
            conn.execute(
                """
                INSERT INTO CMS_FooterContent (id)
                VALUES (1)
                ON CONFLICT (id) DO NOTHING
                """
            )
            footer = conn.execute("SELECT * FROM CMS_FooterContent WHERE id = 1").fetchone()


        def _clean_text(val: str | None) -> str:
            return (val or "").strip()

        def _validate_email(email: str) -> bool:
            email = (email or "").strip()
            if not email:
                return True
            if "@" not in email:
                return False
            return True

        def _validate_phone(phone: str) -> bool:
            phone = (phone or "").strip()
            if not phone:
                return True
            # Allow leading + and common separators/spaces
            for ch in phone:
                if ch.isdigit():
                    continue
                if ch in "+() -.":
                    continue
                return False
            digits = [c for c in phone if c.isdigit()]
            return len(digits) >= 7

        def _validate_url(url: str) -> bool:
            url = (url or "").strip()
            if not url:
                return True
            return url.startswith("http://") or url.startswith("https://")

        if request.method == "POST":
            business_name = _clean_text(request.form.get("business_name"))
            business_tagline = _clean_text(request.form.get("business_tagline"))

            email = _clean_text(request.form.get("email"))
            phone = _clean_text(request.form.get("phone"))
            secondary_phone = _clean_text(request.form.get("secondary_phone"))
            mobile = _clean_text(request.form.get("mobile"))
            whatsapp_number = _clean_text(request.form.get("whatsapp_number"))

            office_address = _clean_text(request.form.get("office_address"))
            google_maps_link = _clean_text(request.form.get("google_maps_link"))
            business_hours = (request.form.get("business_hours") or "").strip()

            facebook_url = _clean_text(request.form.get("facebook_url"))
            instagram_url = _clean_text(request.form.get("instagram_url"))
            x_url = _clean_text(request.form.get("x_url"))
            tiktok_url = _clean_text(request.form.get("tiktok_url"))
            footer_description = (request.form.get("footer_description") or "").strip()
            contact_section_title = (request.form.get("contact_section_title") or "").strip()
            copyright_text = (request.form.get("copyright_text") or "").strip()

            # Validate
            errors = []
            if not _validate_email(email):
                errors.append("Invalid email format.")
            if not _validate_phone(phone):
                errors.append("Invalid primary phone number.")
            if not _validate_phone(secondary_phone):
                errors.append("Invalid secondary phone number.")
            if not _validate_phone(mobile):
                errors.append("Invalid mobile number.")
            if not _validate_phone(whatsapp_number):
                errors.append("Invalid WhatsApp number.")

            for label, u in [
                ("Google Maps URL", google_maps_link),
                ("Facebook URL", facebook_url),
                ("Instagram URL", instagram_url),
                ("X URL", x_url),
                ("TikTok URL", tiktok_url),
            ]:
                if not _validate_url(u):
                    errors.append(f"Invalid {label}. Must start with http:// or https://")

            # Footer logo is no longer edited from this admin flow.
            footer_logo_path = footer["footer_logo_path"] if footer else ""


            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template(
                    "admin_contact_info.html",
                    contact=contact,
                    footer=footer,
                ), 400

            # Persist CMS_ContactInfo
            conn.execute(
                """
                UPDATE CMS_ContactInfo SET
                    business_name = ?,
                    business_tagline = ?,
                    phone = ?,
                    secondary_phone = ?,
                    mobile = ?,
                    whatsapp_number = ?,
                    email = ?,
                    office_address = ?,
                    google_maps_link = ?,
                    business_hours = ?,
                    facebook_url = ?,
                    instagram_url = ?,
                    x_url = ?,
                    tiktok_url = ?
                WHERE id = 1
                """,
                (
                    business_name,
                    business_tagline,
                    phone,
                    secondary_phone,
                    mobile,
                    whatsapp_number,
                    email,
                    office_address,
                    google_maps_link,
                    business_hours,
                    facebook_url,
                    instagram_url,
                    x_url,
                    tiktok_url,
                ),
            )

            # Persist CMS_FooterContent
            conn.execute(
                """
                UPDATE CMS_FooterContent SET
                    footer_description = ?,
                    contact_section_title = ?,
                    copyright_text = ?
                WHERE id = 1
                """,
                (
                    footer_description,
                    contact_section_title,
                    copyright_text,
                ),
            )

            # Keep CMS_SocialLinks in sync with the platform columns.
            # Platform mapping: facebook/instagram/x/tiktok
            conn.execute("DELETE FROM CMS_SocialLinks WHERE platform IN ('facebook','instagram','x','tiktok')")
            for i, (platform, href) in enumerate([
                ("facebook", facebook_url),
                ("instagram", instagram_url),
                ("x", x_url),
                ("tiktok", tiktok_url),
            ]):
                href = (href or "").strip()
                if href:
                    conn.execute(
                        "INSERT INTO CMS_SocialLinks (platform, href, is_visible, link_order) VALUES (?, ?, 1, ?)",
                        (platform, href, i),
                    )


            conn.commit()
            flash("Contact information saved.", "success")
            return redirect(url_for("admin_contact_info"))

        return render_template(
            "admin_contact_info.html",
            contact=contact,
            footer=footer,
        )

    except Exception as e:
        logger.error(f"Admin contact info error: {e}", exc_info=True)
        flash("Error loading/saving contact information.", "error")
        # best-effort render
        return render_template(
            "admin_contact_info.html",
            contact=None,
            footer=None,
        ), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/admin/settings", methods=["GET", "POST"])
@admin_login_required
def admin_settings():

    admin_id = session.get("admin_id")
    conn = get_db_connection()
    admin = conn.execute("SELECT * FROM Admin WHERE id = ?", (admin_id,)).fetchone()

    if request.method == "POST":
        new_username = (request.form.get("username", "") or "").strip()
        new_email = (request.form.get("email", "") or "").strip()
        current_password = (request.form.get("current_password", "") or "").strip()
        new_password = (request.form.get("new_password", "") or "").strip()
        confirm_password = (request.form.get("confirm_password", "") or "").strip()

        if not admin or not check_password_hash(admin["password"], current_password):
            conn.close()
            flash("Current password is incorrect.", "error")
            return redirect(url_for("admin_settings"))

        email_changed = _admin_email_changed(conn, admin_id, new_email)
        email_enabled_value = 0 if email_changed else (admin["email_enabled"] or 0)

        if new_password:
            if len(new_password) < 6:
                conn.close()
                flash("New password must be at least 6 characters.", "error")
                return redirect(url_for("admin_settings"))
            if new_password != confirm_password:
                conn.close()
                flash("New password confirmation does not match.", "error")
                return redirect(url_for("admin_settings"))
            password_hash = generate_password_hash(new_password)
            conn.execute(
                "UPDATE Admin SET username = ?, email = ?, password = ?, email_enabled = ? WHERE id = ?",
                (new_username or admin["username"], new_email, password_hash, email_enabled_value, admin_id),
            )
        else:
            conn.execute(
                "UPDATE Admin SET username = ?, email = ?, email_enabled = ? WHERE id = ?",
                (new_username or admin["username"], new_email, email_enabled_value, admin_id),
            )

        conn.commit()
        conn.close()
        session["admin_username"] = new_username or session.get("admin_username")
        if email_changed:
            flash("Admin settings updated. Email was changed, so email notifications have been disabled. Please activate them manually.", "info")
        else:
            flash("Admin settings updated.", "success")
        return redirect(url_for("admin_settings"))

    conn.close()
    return render_template("admin_settings.html", admin=admin)


@app.route("/admin/toggle-email", methods=["POST"])
@admin_login_required
def admin_toggle_email():
    admin_id = session.get("admin_id")
    conn = get_db_connection()
    admin = conn.execute("SELECT email, email_enabled FROM Admin WHERE id = ?", (admin_id,)).fetchone()
    if not admin:
        conn.close()
        flash("Admin not found.", "error")
        return redirect(url_for("admin_settings"))

    if not admin["email"]:
        conn.close()
        flash("Please set an admin email before activating notifications.", "error")
        return redirect(url_for("admin_settings"))

    new_status = 0 if (admin["email_enabled"] or 0) == 1 else 1
    conn.execute("UPDATE Admin SET email_enabled = ? WHERE id = ?", (new_status, admin_id))
    conn.commit()
    conn.close()
    flash(
        "Email notifications enabled." if new_status == 1 else "Email notifications disabled.",
        "success",
    )
    return redirect(url_for("admin_settings"))


@app.route("/admin/test-email", methods=["POST"])
@admin_login_required
def admin_test_email():
    ok, msg = send_admin_notification("Test Email", "This is a test email from your BusResort booking system.")
    flash(msg, "success" if ok else "error")
    return redirect(url_for("admin_settings"))


# -- Sales Reports --

def _build_date_where_clause(start_date, end_date):
    """Return SQL WHERE clause snippet and params for created_at date filtering."""
    if start_date and end_date:
        return " AND DATE(created_at) BETWEEN ? AND ?", (start_date, end_date)
    elif start_date:
        return " AND DATE(created_at) >= ?", (start_date,)
    elif end_date:
        return " AND DATE(created_at) <= ?", (end_date,)
    return "", ()


@app.route("/admin/sales")
@admin_login_required
def admin_sales():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where_bus, date_params_bus = _build_date_where_clause(start_date, end_date)
    date_where_resort, date_params_resort = _build_date_where_clause(start_date, end_date)

    bus = conn.execute(
        f"SELECT price, created_at, status FROM BusBookings WHERE 1=1{date_where_bus}",
        date_params_bus
    ).fetchall()
    resort = conn.execute(
        f"SELECT price, created_at, status FROM ResortBookings WHERE 1=1{date_where_resort}",
        date_params_resort
    ).fetchall()
    conn.close()

    def _day_key(ts: str) -> str:
        return (ts or "")[:10] or "Unknown"

    daily: dict[str, float] = {}
    monthly: dict[str, float] = {}
    total = 0.0
    bus_total = 0.0
    resort_total = 0.0
    bus_count = 0
    resort_count = 0
    
    # Status counts for complete overview
    bus_confirmed = bus_pending = bus_cancelled = 0
    resort_confirmed = resort_pending = resort_cancelled = 0

    for row in bus:
        status = row["status"] or "Pending"
        price = float(row["price"] or 0)
        
        if status == "Confirmed":
            bus_confirmed += 1
            bus_total += price
            total += price
            bus_count += 1
            d = _day_key(row["created_at"])
            m = d[:7] if len(d) >= 7 else "Unknown"
            daily[d] = daily.get(d, 0.0) + price
            monthly[m] = monthly.get(m, 0.0) + price
        elif status == "Pending":
            bus_pending += 1
        elif status == "Cancelled":
            bus_cancelled += 1

    for row in resort:
        status = row["status"] or "Pending"
        price = float(row["price"] or 0)
        
        if status == "Confirmed":
            resort_confirmed += 1
            resort_total += price
            total += price
            resort_count += 1
            d = _day_key(row["created_at"])
            m = d[:7] if len(d) >= 7 else "Unknown"
            daily[d] = daily.get(d, 0.0) + price
            monthly[m] = monthly.get(m, 0.0) + price
        elif status == "Pending":
            resort_pending += 1
        elif status == "Cancelled":
            resort_cancelled += 1

    daily_items = sorted(daily.items(), key=lambda x: x[0], reverse=True)
    monthly_items = sorted(monthly.items(), key=lambda x: x[0], reverse=True)
    
    return render_template(
        "sales.html",
        daily_items=daily_items,
        monthly_items=monthly_items,
        total=total,
        bus_total=bus_total,
        resort_total=resort_total,
        bus_count=bus_count,
        resort_count=resort_count,
        bus_confirmed=bus_confirmed,
        bus_pending=bus_pending,
        bus_cancelled=bus_cancelled,
        resort_confirmed=resort_confirmed,
        resort_pending=resort_pending,
        resort_cancelled=resort_cancelled,
        filter_type=filter_type,
        start_date=start_date_str,
        end_date=end_date_str,
        period_label=period_label,
    )


@app.route("/admin/sales/bus")
@admin_login_required
def bus_sales():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    total = sum(float(r["price"] or 0) for r in rows)
    conn.close()
    return render_template(
        "bus_sales.html",
        sales=rows,
        total=total,
        filter_type=filter_type,
        start_date=start_date_str,
        end_date=end_date_str,
        period_label=period_label,
    )


@app.route("/admin/sales/resort")
@admin_login_required
def resort_sales():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    total = sum(float(r["total_cost"] or r["price"] or 0) for r in rows)
    conn.close()
    return render_template(
        "resort_sales.html",
        sales=rows,
        total=total,
        filter_type=filter_type,
        start_date=start_date_str,
        end_date=end_date_str,
        period_label=period_label,
    )


# -- Excel Export (openpyxl) --

def _create_excel_workbook(title: str, period_label: str, total: float, entries: int):
    """Create a styled Excel workbook with summary header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.append([title])
    ws.append([f"Period: {period_label or 'All Time'}"])
    ws.append([f"Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Total Revenue: PHP {total:,.2f}"])
    ws.append([f"Total Transactions: {entries}"])
    ws.append([])

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=1):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True, size=16)
            elif cell.row in (4, 5):
                cell.font = summary_font
            cell.alignment = Alignment(horizontal='left')

    return wb, ws, header_font, header_fill, thin_border


@app.route("/admin/export/bus/xlsx")
@admin_login_required
def export_bus_xlsx():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    total = sum(float(r["price"] or 0) for r in rows)
    wb, ws, header_font, header_fill, thin_border = _create_excel_workbook(
        "Bus Sales Report", period_label, total, len(rows)
    )

    headers = ["Booking ID", "Customer Name", "Contact", "Email", "Pickup", "Destination",
               "Travel Date/Time", "Amount (PHP)", "Status", "Booking Date"]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r in rows:
        ws.append([
            r["id"], r["name"], r["contact"] or "", r["email"] or "",
            r["pickup"] or "", r["destination"] or "", r["datetime"] or "",
            float(r["price"] or 0), r["status"] or "",
            (r["created_at"] or "")[:10]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bus_sales.xlsx"
    )


@app.route("/admin/export/resort/xlsx")
@admin_login_required
def export_resort_xlsx():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    total = sum(float(r["total_cost"] or r["price"] or 0) for r in rows)
    wb, ws, header_font, header_fill, thin_border = _create_excel_workbook(
        "Resort Sales Report", period_label, total, len(rows)
    )

    headers = ["Booking ID", "Customer Name", "Contact", "Email", "Room Type",
               "Check-in", "Check-out", "Check-in Time", "Check-out Time", "Guests",
               "Payment Method", "Price/Night (PHP)", "Total Cost (PHP)",
               "Exclusive", "Appliances Cost (PHP)", "Amount (PHP)", "Status", "Booking Date"]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r in rows:
        exclusive = "Yes" if (r["is_exclusive"] or 0) == 1 else "No"
        ws.append([
            r["id"], r["name"], r["contact"] or "", r["email"] or "",
            r["room_type"] or "", r["checkin"] or "", r["checkout"] or "",
            r["checkin_time"] or "14:00", r["checkout_time"] or "12:00",
            r["guests"] or 0, r["payment_method"] or "Cash",
            float(r["price_per_night"] or 0), float(r["total_cost"] or 0),
            exclusive, float(r["appliances_cost"] or 0), float(r["price"] or 0),
            r["status"] or "", (r["created_at"] or "")[:10]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resort_sales.xlsx"
    )


# -- PDF Export --

def _render_pdf_or_html(template_name: str, **ctx):
    """Render PDF via weasyprint if available, otherwise return HTML for browser print."""
    html_out = render_template(template_name, **ctx)
    try:
        import importlib
        weasyprint_mod = importlib.import_module("weasyprint")
        HTML = weasyprint_mod.HTML
        pdf_bytes = io.BytesIO()
        HTML(string=html_out).write_pdf(pdf_bytes)
        pdf_bytes.seek(0)
        return send_file(
            pdf_bytes,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=ctx.get("filename", "report.pdf")
        )
    except Exception:
        return Response(html_out, mimetype="text/html")


@app.route("/admin/export/bus/pdf")
@admin_login_required
def export_bus_pdf():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    total = sum(float(r["price"] or 0) for r in rows)
    conn.close()

    return _render_pdf_or_html(
        "bus_sales_pdf.html",
        title="Bus Sales Report",
        sales=rows,
        total=total,
        entries=len(rows),
        now=dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        filename="bus_sales.pdf",
        period_label=period_label,
    )


@app.route("/admin/export/resort/pdf")
@admin_login_required
def export_resort_pdf():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    total = sum(float(r["total_cost"] or r["price"] or 0) for r in rows)
    conn.close()

    return _render_pdf_or_html(
        "resort_sales_pdf.html",
        title="Resort Sales Report",
        sales=rows,
        total=total,
        entries=len(rows),
        now=dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        filename="resort_sales.pdf",
        period_label=period_label,
    )


# -- Combined Export --

@app.route("/admin/export/combined/xlsx")
@admin_login_required
def export_combined_xlsx():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)

    bus_rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    resort_rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    bus_total = sum(float(r["price"] or 0) for r in bus_rows)
    resort_total = sum(float(r["total_cost"] or r["price"] or 0) for r in resort_rows)
    overall_total = bus_total + resort_total

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws_summary.append(["BusResort Combined Sales Report"])
    ws_summary.append([f"Period: {period_label or 'All Time'}"])
    ws_summary.append([f"Generated: {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_summary.append([])
    ws_summary.append(["Category", "Transactions", "Total Revenue (PHP)"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws_summary.append(["Bus Sales", len(bus_rows), f"{bus_total:,.2f}"])
    ws_summary.append(["Resort Sales", len(resort_rows), f"{resort_total:,.2f}"])
    ws_summary.append(["Overall Total", len(bus_rows) + len(resort_rows), f"{overall_total:,.2f}"])

    for row in ws_summary.iter_rows(min_row=1, max_row=3, min_col=1, max_col=1):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal='left')

    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_summary.column_dimensions[column_letter].width = min(max_length + 2, 50)

    ws_bus = wb.create_sheet(title="Bus Sales")
    bus_headers = ["Booking ID", "Customer Name", "Contact", "Email", "Pickup", "Destination",
                   "Travel Date/Time", "Amount (PHP)", "Status", "Booking Date"]
    ws_bus.append(bus_headers)
    for cell in ws_bus[ws_bus.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r in bus_rows:
        ws_bus.append([
            r["id"], r["name"], r["contact"] or "", r["email"] or "",
            r["pickup"] or "", r["destination"] or "", r["datetime"] or "",
            float(r["price"] or 0), r["status"] or "",
            (r["created_at"] or "")[:10]
        ])

    for column in ws_bus.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_bus.column_dimensions[column_letter].width = min(max_length + 2, 50)

    ws_resort = wb.create_sheet(title="Resort Sales")
    resort_headers = ["Booking ID", "Customer Name", "Contact", "Email", "Room Type",
                      "Check-in", "Check-out", "Check-in Time", "Check-out Time", "Guests",
                      "Payment Method", "Price/Night (PHP)", "Total Cost (PHP)",
                      "Exclusive", "Appliances Cost (PHP)", "Amount (PHP)", "Status", "Booking Date"]
    ws_resort.append(resort_headers)
    for cell in ws_resort[ws_resort.max_row]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r in resort_rows:
        exclusive = "Yes" if (r["is_exclusive"] or 0) == 1 else "No"
        ws_resort.append([
            r["id"], r["name"], r["contact"] or "", r["email"] or "",
            r["room_type"] or "", r["checkin"] or "", r["checkout"] or "",
            r["checkin_time"] or "14:00", r["checkout_time"] or "12:00",
            r["guests"] or 0, r["payment_method"] or "Cash",
            float(r["price_per_night"] or 0), float(r["total_cost"] or 0),
            exclusive, float(r["appliances_cost"] or 0), float(r["price"] or 0),
            r["status"] or "", (r["created_at"] or "")[:10]
        ])

    for column in ws_resort.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_resort.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="combined_sales.xlsx"
    )


@app.route("/admin/export/combined/pdf")
@admin_login_required
def export_combined_pdf():
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)

    bus_rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    resort_rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    bus_total = sum(float(r["price"] or 0) for r in bus_rows)
    resort_total = sum(float(r["total_cost"] or r["price"] or 0) for r in resort_rows)
    overall_total = bus_total + resort_total

    return _render_pdf_or_html(
        "combined_sales_pdf.html",
        title="Combined Sales Report",
        bus_sales=bus_rows,
        resort_sales=resort_rows,
        bus_total=bus_total,
        resort_total=resort_total,
        overall_total=overall_total,
        bus_entries=len(bus_rows),
        resort_entries=len(resort_rows),
        total_entries=len(bus_rows) + len(resort_rows),
        now=dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        filename="combined_sales.pdf",
        period_label=period_label,
    )


# -- CSV Export --

@app.route("/admin/export/bus/csv")
@admin_login_required
def export_bus_csv():
    """Export bus sales to CSV format."""
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM BusBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Booking ID", "Customer Name", "Contact", "Email", "Pickup", "Destination",
        "Travel Date/Time", "Amount (PHP)", "Status", "Booking Date", "Notes"
    ])
    
    # Data
    for r in rows:
        writer.writerow([
            r["id"],
            r["name"],
            r["contact"] or "",
            r["email"] or "",
            r["pickup"] or "",
            r["destination"] or "",
            r["datetime"] or "",
            float(r["price"] or 0),
            r["status"] or "",
            (r["created_at"] or "")[:10],
            r["notes"] or ""
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bus_sales_{dt.datetime.now().strftime('%Y%m%d')}.csv"}
    )


@app.route("/admin/export/resort/csv")
@admin_login_required
def export_resort_csv():
    """Export resort sales to CSV format."""
    filter_type = request.args.get("filter", "").strip()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    start_date, end_date, period_label = _get_date_range(filter_type, start_date_str, end_date_str)

    conn = get_db_connection()
    date_where, date_params = _build_date_where_clause(start_date, end_date)
    rows = conn.execute(
        f"SELECT * FROM ResortBookings WHERE status = 'Confirmed'{date_where} ORDER BY id DESC",
        date_params
    ).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Booking ID", "Customer Name", "Contact", "Email", "Room Type",
        "Check-in", "Check-out", "Check-in Time", "Check-out Time", "Guests",
        "Payment Method", "Price/Night (PHP)", "Total Cost (PHP)",
        "Exclusive", "Appliances Cost (PHP)", "Amount (PHP)", "Status", "Booking Date", "Notes"
    ])
    
    # Data
    for r in rows:
        exclusive = "Yes" if (r["is_exclusive"] or 0) == 1 else "No"
        writer.writerow([
            r["id"],
            r["name"],
            r["contact"] or "",
            r["email"] or "",
            r["room_type"] or "",
            r["checkin"] or "",
            r["checkout"] or "",
            r["checkin_time"] or "14:00",
            r["checkout_time"] or "12:00",
            r["guests"] or 0,
            r["payment_method"] or "Cash",
            float(r["price_per_night"] or 0),
            float(r["total_cost"] or 0),
            exclusive,
            float(r["appliances_cost"] or 0),
            float(r["price"] or 0),
            r["status"] or "",
            (r["created_at"] or "")[:10],
            r["notes"] or ""
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=resort_sales_{dt.datetime.now().strftime('%Y%m%d')}.csv"}
    )


@app.route("/admin/email/send", methods=["POST"])
@admin_login_required
def admin_send_email():
    to_email = (request.form.get("to_email", "") or "").strip()
    subject = (request.form.get("subject", "") or "").strip() or "BusResort Message"
    message = (request.form.get("message", "") or "").strip()
    ok, msg = _send_email(to_email, subject, message)
    flash(msg if ok else f"Email failed: {msg}", "success" if ok else "error")
    return redirect(url_for("dashboard"))


@app.route("/api/admin/calendar_events")
@admin_login_required
def admin_calendar_events():
    """Return calendar events for a bounded date window.

    Root cause of intermittent hangs: this endpoint previously loaded ALL bookings
    and built a potentially huge events array on every dashboard load.

    Now we default to a small window around today and allow the frontend to
    request a custom range via query params:
      - start=YYYY-MM-DD
      - end=YYYY-MM-DD

    If dates are missing/invalid, fallback to [today-30d, today+60d].
    """
    conn = None
    try:
        conn = get_db_connection(timeout=10.0)

        start_date = request.args.get('start', '').strip()
        end_date = request.args.get('end', '').strip()

        # Parse ISO dates safely (YYYY-MM-DD)
        def _parse_iso_date(s: str):
            try:
                return dt.date.fromisoformat(s)
            except Exception:
                return None

        today = dt.date.today()
        s_date = _parse_iso_date(start_date)
        e_date = _parse_iso_date(end_date)

        # Default window: last 30 days + next 60 days
        if not s_date:
            s_date = today - dt.timedelta(days=30)
        if not e_date:
            e_date = today + dt.timedelta(days=60)
        if s_date > e_date:
            s_date, e_date = e_date, s_date

        s_str = s_date.strftime('%Y-%m-%d')
        e_str = e_date.strftime('%Y-%m-%d')

        # Only load bookings that overlap the requested window.
        # BusBookings uses `datetime` (and checkout as end marker).
        bus = conn.execute(
            """
            SELECT id, name, datetime, checkin, checkout, status
            FROM BusBookings
            WHERE status IN ('Pending','Confirmed')
              AND (
                    (checkout IS NOT NULL AND DATE(checkout) >= DATE(?))
                    OR (datetime IS NOT NULL AND DATE(datetime) <= DATE(?) AND DATE(datetime) >= DATE(?))
                    OR (datetime IS NOT NULL AND DATE(datetime) <= DATE(?))
                  )
            ORDER BY id DESC
            """,
            (s_str, e_str, s_str, e_str),
        ).fetchall()

        # ResortBookings uses `checkin` and `checkout` (with overlap logic).
        resort = conn.execute(
            """
            SELECT id, name, checkin, checkout, checkin_time, checkout_time, status
            FROM ResortBookings
            WHERE status IN ('Pending','Confirmed')
              AND (
                    (checkout IS NOT NULL AND DATE(checkout) >= DATE(?))
                    OR (checkin IS NOT NULL AND DATE(checkin) <= DATE(?) AND DATE(checkin) >= DATE(?))
                    OR (checkin IS NOT NULL AND DATE(checkin) <= DATE(?))
                  )
            ORDER BY id DESC
            """,
            (s_str, e_str, s_str, e_str),
        ).fetchall()

        logger.info(
            f"[Calendar API] window={s_str}..{e_str} raw data: {len(bus)} bus, {len(resort)} resort"
        )

        events = []

        # Build bus events
        for b in bus:
            try:
                start_raw = _row_value(b, "datetime", "") or _row_value(b, "checkin", "")
                start = str(start_raw).strip() if start_raw else ""
                if not start:
                    start = today.strftime('%Y-%m-%d')

                end_raw = _row_value(b, "checkout", "") or start
                end = str(end_raw).strip() if end_raw else start

                status = _row_value(b, "status", "Pending")
                name = _row_value(b, "name", "Unknown")

                events.append({
                    "title": f"Bus: {name} ({status})",
                    "start": start,
                    "end": end or start,
                    "allDay": False,
                    "url": f"/edit/bus/{b['id']}",
                    "color": "#22c55e" if status == "Confirmed" else "#f59e0b",
                    "extendedProps": {
                        "type": "bus",
                        "booking_id": b['id'],
                        "status": status,
                    },
                })
            except Exception as row_err:
                logger.warning(f"[Calendar API] Error bus booking {b.get('id','unknown')}: {row_err}")
                continue

        # Build resort events
        for r in resort:
            try:
                start_raw = _row_value(r, "checkin", "")
                start = str(start_raw).strip() if start_raw else ""
                if not start:
                    start = today.strftime('%Y-%m-%d')

                checkin_time = str(_row_value(r, "checkin_time", "14:00")).strip()
                checkout_time = str(_row_value(r, "checkout_time", "12:00")).strip()

                if "T" not in start and checkin_time:
                    start = f"{start}T{checkin_time}"

                end_raw = _row_value(r, "checkout", "")
                end = str(end_raw).strip() if end_raw else ""
                if end and checkout_time and "T" not in end:
                    end = f"{end}T{checkout_time}"

                status = _row_value(r, "status", "Pending")
                name = _row_value(r, "name", "Unknown")

                events.append({
                    "title": f"Resort: {name} ({status})",
                    "start": start,
                    "end": end or start,
                    "allDay": False,
                    "url": f"/edit/resort/{r['id']}",
                    "color": "#3b82f6" if status == "Confirmed" else "#f59e0b",
                    "extendedProps": {
                        "type": "resort",
                        "booking_id": r['id'],
                        "status": status,
                    },
                })
            except Exception as row_err:
                logger.warning(f"[Calendar API] Error resort booking {r.get('id','unknown')}: {row_err}")
                continue

        logger.info(f"[Calendar API] total events: {len(events)}")
        return jsonify(events)

    except Exception as e:
        logger.error(f"[Calendar API] Fatal error: {e}", exc_info=True)
        return jsonify({"error": str(e), "events": []}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass



@app.route("/admin/feedback")
@admin_login_required
def admin_feedback_view():
    conn = get_db_connection()
    feedback_list = conn.execute("SELECT * FROM Feedback ORDER BY date_created DESC").fetchall()
    conn.close()
    return render_template("admin_feedback.html", feedback=feedback_list)


@app.route("/admin/feedback/delete/<int:feedback_id>", methods=["POST"])
@admin_login_required
def admin_delete_feedback(feedback_id: int):
    conn = get_db_connection()
    result = conn.execute("DELETE FROM Feedback WHERE id = ?", (feedback_id,)).rowcount
    conn.commit()
    conn.close()
    if result:
        flash("Feedback deleted successfully.", "success")
    else:
        flash("Feedback not found.", "error")
    return redirect(url_for("admin_feedback_view"))


@app.route("/admin/notifications/mark-read/<int:notif_id>", methods=["POST"])
@admin_login_required
def admin_mark_notification_read(notif_id):
    conn = get_db_connection()
    conn.execute("UPDATE Notification SET is_read = 1 WHERE id = ?", (notif_id,))
    conn.commit()
    conn.close()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({"success": True, "message": "Notification marked as read"})
    flash("Notification marked as read.", "success")
    return redirect(url_for("dashboard"))


@app.route("/api/admin/notifications")
@admin_login_required
def api_admin_notifications():
    conn = None
    try:
        conn = get_db_connection(timeout=5.0)  # Short timeout for API
        unread_count = conn.execute("SELECT COUNT(*) FROM Notification WHERE is_read = 0").fetchone()[0]
        notifications = conn.execute("SELECT * FROM Notification ORDER BY date_created DESC LIMIT 50").fetchall()
        return jsonify({
            "unread_count": unread_count,
            "notifications": [dict(n) for n in notifications]
        })
    except Exception as e:
        logger.error(f"Notifications API error: {e}", exc_info=True)
        return jsonify({"unread_count": 0, "notifications": [], "error": "Failed to load notifications"}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/api/admin/notifications/<int:notif_id>", methods=["DELETE"])
@admin_login_required
def api_admin_delete_notification(notif_id: int):
    conn = None
    try:
        conn = get_db_connection(timeout=5.0)
        conn.execute("DELETE FROM Notification WHERE id = ?", (notif_id,))
        conn.commit()
        return jsonify({"success": True, "message": "Notification deleted"})
    except Exception as e:
        logger.error(f"Delete notification error: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Failed to delete notification"}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@app.route("/api/admin/notifications/mark-all-read", methods=["POST"])
@admin_login_required
def api_admin_mark_all_read():
    conn = None
    try:
        conn = get_db_connection(timeout=5.0)
        conn.execute("UPDATE Notification SET is_read = 1 WHERE is_read = 0")
        conn.commit()
        return jsonify({"success": True, "message": "All notifications marked as read"})
    except Exception as e:
        logger.error(f"Mark all read error: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Failed to mark notifications as read"}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


# Serve uploads from shared project uploads folder
@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    """Serve uploaded files from the project's static/uploads folder."""
    upload_folder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'uploads')
    return send_from_directory(upload_folder, filename)


if __name__ == '__main__':
    # Initialize database and ensure WebsiteSettings table exists
    init_db()
    init_website_settings()
    app.run(debug=False, host="127.0.0.1", port=5001)
